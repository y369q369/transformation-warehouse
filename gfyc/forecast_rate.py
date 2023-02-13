import argparse
import datetime
import math
import sys
from decimal import *

import influxdb_client
import pymysql
import pytz
import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side

parser = argparse.ArgumentParser(description='统计每日光伏功率预测准确率，默认查询昨天')
parser.add_argument('--type', type=int, help='1：区县  2：馈线  3：台区  4：电表', default=1)
parser.add_argument('--queryId', type=str, help='查询编号', default='3240108')
parser.add_argument('--startDate', type=str, help='开始日期')
parser.add_argument('--endDate', type=str, help='结束日期')

target_file = './forecast_power_rate{}.xlsx'
start_time_str = '08:00:00'
end_time_str = '17:00:00'

date_list = []
calculation_data = {}
target_data = []

conn = pymysql.connect(user='root', password='123456789', host='172.16.130.188', database='gfyc')
cursor = conn.cursor()

client = influxdb_client.InfluxDBClient(url='http://172.16.130.205:8092/',
                                        token='LRdIIW17oir2cDvCsrjZn1qvUOF5fFNwlqeqHGvg5LAv7pw-g_efZNbp7hhvV1aZwBecmMNgeE8dO9yPIPdLqA==',
                                        org='nari')
query_api = client.query_api()


# 本地时间转utc时间
def local_utc(datetime_str):
    local_time = datetime.datetime.strptime(datetime_str, '%Y-%m-%d %H:%M:%S')
    utc_time = local_time.astimezone(pytz.timezone('UTC'))
    utc_time_str = datetime.datetime.strftime(utc_time, '%Y-%m-%dT%H:%M:%SZ')
    return utc_time_str


# 获取预测值
def get_forecast(field, id, county_code):
    curve_data = {}
    for date_str in date_list:
        curve_data[date_str] = {'power': {}}
        forecast_sql = 'SELECT \n' \
                       '    b.time, round(sum( a.per * b.FORECAST ), 2) \n' \
                       'FROM \n' \
                       '    pv_device_per_full a, gfyc_shortforecast b \n' \
                       'WHERE \n' \
                       '    a.county_code = b.id \n' \
                       '    AND b.time BETWEEN "{} {}" AND "{} {}" \n' \
                       '    AND a.county_code = "{}" \n' \
                       '    AND a.{} = "{}" \n' \
                       '    GROUP BY b.time \n'.format(date_str, start_time_str, date_str, end_time_str, county_code,
                                                       field, id)
        print(forecast_sql)
        cursor.execute(forecast_sql)
        forecast_result = cursor.fetchall()
        for item in forecast_result:
            time_str = datetime.datetime.strftime(item[0], '%Y-%m-%d %H:%M')
            curve_data[date_str]['power'][time_str] = {
                'forecast': item[1]
            }
    calculation_data['date'] = curve_data


# 获取实际值
def get_actual(type, id, county_code):
    for date_str in date_list:
        if type == 1:
            filter_str = ''
        elif type == 2:
            platform_id_list_sql = 'select distinct resrc_supl_code from dwd_cst_dygfyhsb_df where line_id = "{}" \n'.format(
                id)
            print(platform_id_list_sql)
            cursor.execute(platform_id_list_sql)
            platform_id_list_result = cursor.fetchall()
            filter_str = '  |> filter(fn: (r) =>'
            for i in range(0, len(platform_id_list_result)):
                if i > 0:
                    filter_str += ' or'
                filter_str += ' r["platform_id"] == "{}"'.format(platform_id_list_result[i][0])
            filter_str += ')\n'
        elif type == 3:
            filter_str = '  |> filter(fn: (r) => r["platform_id"] == "{}")\n'.format(id)
        elif type == 4:
            filter_str = '  |> filter(fn: (r) => r["meter_id"] == "{}")\n'.format(id)
        actual_query = 'from(bucket: "{}")\n  |> range(start: {}, stop: {})\n' \
                       '{}' \
                       '  |> group(columns: ["_time"])\n' \
                       '  |> sum()\n' \
            .format(county_code, local_utc(date_str + ' ' + start_time_str), local_utc(date_str + ' ' + end_time_str),
                    filter_str)
        print(actual_query)
        actual_result = query_api.query(org='nari', query=actual_query)
        for item in actual_result:
            for info in item.records:
                time_str = datetime.datetime.strftime(info['_time'].astimezone(pytz.timezone('Asia/Shanghai')),
                                                      '%Y-%m-%d %H:%M')
                if time_str in calculation_data['date'][date_str]['power']:
                    calculation_data['date'][date_str]['power'][time_str]['actual'] = round(info['_value'] * 1000, 2)


# 获取额定功率
def get_inst_cap(field, id):
    inst_cap_sql = 'select sum(inst_cap * 1000) from dwd_inst_cap where {} = "{}" \n'.format(field, id)
    print(inst_cap_sql)
    cursor.execute(inst_cap_sql)
    inst_cap_result = cursor.fetchone()
    calculation_data['inst_cap'] = inst_cap_result[0]


# 获取天气
def get_weather(county_code):
    for date_str in date_list:
        weather_sql = 'SELECT \n' \
                      '     CASE \n' \
                      '         s.weather_text \n' \
                      '         WHEN "晴" THEN "晴" \n' \
                      '         WHEN "多云" THEN "多云" \n' \
                      '         ELSE "阴雨" \n' \
                      '     END weather, \n' \
                      '     COUNT( 0 ) num  \n' \
                      'FROM \n' \
                      '     gfyc_relation g, \n' \
                      '     system_forecast_qx_hour_buffer_all s \n' \
                      'WHERE \n' \
                      '     g.station_id = s.id \n' \
                      '     AND g.county_code = "{}" \n' \
                      '     AND s.yb_time between "{} {}" and "{} {}" \n' \
                      'GROUP BY weather \n' \
                      'ORDER BY num DESC \n' \
                      'LIMIT 1 \n'.format(county_code, date_str, start_time_str, date_str, end_time_str)
        print(weather_sql)
        cursor.execute(weather_sql)
        inst_cap_result = cursor.fetchone()
        calculation_data['date'][date_str]['weather'] = inst_cap_result[0]


# 计算准确率
def calculation_percent():
    if calculation_data['inst_cap'] != '':
        for date_str in calculation_data['date']:
            weather = calculation_data['date'][date_str]['weather']
            power = calculation_data['date'][date_str]['power']
            powerSquare = Decimal(0)
            num = 0
            for time_str in power:
                if 'actual' in power[time_str] and 'forecast' in power[time_str]:
                    temp_rate = (Decimal(power[time_str]['actual']) - Decimal(power[time_str]['forecast'])) / Decimal(
                        calculation_data['inst_cap'])
                    powerSquare = temp_rate ** 2 + powerSquare
                    num += 1
            if num > 0:
                rate = 1 - math.sqrt(powerSquare / Decimal(num))
                target_data.append([date_str, weather, round(rate * 100, 2)])


# 表头
table_header = ['日期', '天气', '准确率（%）']


# 获取数据
def get_data(type, id):
    query_county_code_sql = 'select distinct county_code from dwd_cst_dygfyhsb_df where {} = "{}"\n'
    query_county_code_field = ''
    global target_file
    if type == 1:
        query_county_code_field = 'county_code'
        target_file = target_file.format('（区县_' + id + '）')
    elif type == 2:
        query_county_code_field = 'line_id'
        target_file = target_file.format('（馈线_' + id + '）')
    elif type == 3:
        query_county_code_field = 'resrc_supl_code'
        target_file = target_file.format('（台区_' + id + '）')
    elif type == 4:
        query_county_code_field = 'meter_id'
        target_file = target_file.format('（电表_' + id + '）')
    else:
        print('类型异常')
        sys.exit()
    query_county_code_sql = query_county_code_sql.format(query_county_code_field, id)
    print(query_county_code_sql)
    cursor.execute(query_county_code_sql)
    county_code_result = cursor.fetchone()
    county_code = county_code_result[0]
    print('区县编码： ' + county_code + '\n')
    # 计算数据
    get_forecast(query_county_code_field, id, county_code)
    get_actual(type, id, county_code)
    get_inst_cap(query_county_code_field, id)
    get_weather(county_code)
    # print(calculation_data)
    calculation_percent()
    print(target_data)
    export_data()


# 数据导出到表格
def export_data():
    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = '预测准确率'
    target_data.insert(0, table_header)
    # 写数据
    row_num = 1
    for row_data in target_data:
        col_num = 1
        for single_data in row_data:
            sheet.cell(row_num, col_num, single_data)
            col_num += 1
        row_num += 1
    # 修改单元格格式
    fill = PatternFill(patternType='solid', start_color='38B0DE')
    font = Font(bold=True)
    thin = Side(border_style='thin', color='000000')
    border = Border(left=thin, right=thin, bottom=thin, top=thin)
    for i in sheet[1]:
        i.fill = fill
        i.font = font
        i.border = border
        sheet.column_dimensions[i.column_letter].width = 40
    book.save(target_file)


if __name__ == '__main__':
    args = parser.parse_args()
    print(args)
    if args.startDate is None or args.endDate is None:
        date_list.append(datetime.datetime.strftime(datetime.datetime.now() + datetime.timedelta(days=-1), '%Y-%m-%d'))
    else:
        startDate = datetime.datetime.strptime(args.startDate, '%Y-%m-%d')
        endDate = datetime.datetime.strptime(args.endDate, '%Y-%m-%d')
        while endDate >= startDate:
            date_list.append(datetime.datetime.strftime(startDate, '%Y-%m-%d'))
            startDate = startDate + datetime.timedelta(days=1)
    get_data(args.type, args.queryId)
