import math

import pymysql
import datetime
import influxdb_client
from decimal import *
import pytz
import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side

client = influxdb_client.InfluxDBClient(url='http://172.16.130.205:8092/',
                                        token='LRdIIW17oir2cDvCsrjZn1qvUOF5fFNwlqeqHGvg5LAv7pw-g_efZNbp7hhvV1aZwBecmMNgeE8dO9yPIPdLqA==',
                                        org='nari')

conn = pymysql.connect(user='root', password='123456789', host='172.16.130.188', database='gfyc')
cursor = conn.cursor()

st = '2023-02-09 08:00:00'
et = '2023-02-09 17:00:00'


# 本地时间转utc时间
def local_utc(datetime_str):
    local_time = datetime.datetime.strptime(datetime_str, '%Y-%m-%d %H:%M:%S')
    utc_time = local_time.astimezone(pytz.timezone('UTC'))
    utc_time_str = datetime.datetime.strftime(utc_time, '%Y-%m-%dT%H:%M:%SZ')
    return utc_time_str


# 导出excel表格
def output_excel_xlsx(filename, table_header, data, sheet_list):
    book = openpyxl.Workbook()

    for i in range(0, len(sheet_list)):
        sheet = book.create_sheet(sheet_list[i], i)
        data[i].insert(0, table_header)
        # 写数据
        row_num = 1
        for row_data in data[i]:
            col_num = 1
            for single_data in row_data:
                sheet.cell(row_num, col_num, single_data)
                col_num += 1
            row_num += 1
        # 修改单元格格式
        fill = PatternFill(patternType='solid', start_color='38B0DE')
        font = Font(bold=True)
        thin = Side(border_style='thin', color='000000')
        border = Border(left=thin, right=thin, bottom=thin, top=thin)
        for i in sheet[1]:
            i.fill = fill
            i.font = font
            i.border = border
            sheet.column_dimensions[i.column_letter].width = 40
    book.save(filename)


# 电表信息
device_dict = {}
# 台区信息
platform_dict = {}
# 馈线信息
line_dict = {}
# 区县信息
county_dict = {}
# 市级信息
city_dict = {}

# 获取预测值与功率
forecast_sql = 'SELECT \n' \
               '    b.time, a.meter_id,	a.cust_name, a.resrc_supl_code, a.resrc_supl_name, a.line_id, a.line_name, a.county_code, a.county_name, a.city_code, a.city_name, a.per * b.FORECAST \n' \
               'FROM \n' \
               '    pv_device_per_full a, gfyc_shortforecast b \n' \
               'WHERE \n' \
               '    a.county_code = b.id \n' \
               '    AND b.time BETWEEN "{}" AND "{}" \n' \
               '    AND a.line_id = "10DKX-108673" \n'.format(st, et)
print(forecast_sql)
cursor.execute(forecast_sql)
forecast_result = cursor.fetchall()
for item in forecast_result:
    time_str = datetime.datetime.strftime(item[0], '%Y-%m-%d %H:%M')
    # 处理电表信息
    if item[1] not in device_dict:
        device_dict[item[1]] = {
            'name': item[2],
            'power': {}
        }
    device_dict[item[1]]['power'][time_str] = {
        'forecast': item[-1]
    }

    # 处理台区信息
    if item[3] not in platform_dict:
        platform_dict[item[3]] = {
            'name': item[4],
            'power': {},
            'device': []
        }
    if time_str not in platform_dict[item[3]]['power']:
        platform_dict[item[3]]['power'][time_str] = {
            'forecast': item[-1],
        }
    else:
        platform_dict[item[3]]['power'][time_str]['forecast'] = float(
            Decimal(item[-1]) + Decimal(platform_dict[item[3]]['power'][time_str]['forecast']))

    # 处理馈线信息
    if item[5] not in line_dict:
        line_dict[item[5]] = {
            'name': item[6],
            'power': {}
        }
    if time_str not in line_dict[item[5]]:
        line_dict[item[5]]['power'][time_str] = {
            'forecast': item[-1]
        }
    else:
        line_dict[item[5]]['power'][time_str]['forecast'] = float(
            Decimal(item[-1]) + Decimal(line_dict[item[5]]['power'][time_str]['forecast']))
    # 处理区县信息
    if item[7] not in county_dict:
        county_dict[item[7]] = {
            'name': item[8],
            'power': {}
        }
    if time_str not in county_dict[item[7]]:
        county_dict[item[7]]['power'][time_str] = {
            'forecast': item[-1]
        }
    else:
        county_dict[item[7]]['power'][time_str]['forecast'] = float(
            Decimal(item[-1]) + Decimal(county_dict[item[7]]['power'][time_str]['forecast']))
    # 处理地市信息
    if item[9] not in city_dict:
        city_dict[item[9]] = {
            'name': item[10],
            'power': {}
        }
    if time_str not in city_dict[item[9]]:
        city_dict[item[9]]['power'][time_str] = {
            'forecast': item[-1]
        }
    else:
        city_dict[item[9]]['power'][time_str]['forecast'] = float(
            Decimal(item[-1]) + Decimal(city_dict[item[9]]['power'][time_str]['forecast']))

# 获取电表功率
forecast_sql = 'select meter_id, inst_cap * 1000 from dwd_cst_dygfyhsb_df'
cursor.execute(forecast_sql)
accs_result = cursor.fetchall()
for item in accs_result:
    if item[0] in device_dict:
        if item[1] != 0:
            device_dict[item[0]]['inst_cap'] = item[1]

# 获取实际值数据
query_api = client.query_api()
for county_code in county_dict.keys():
    infludb_device_query = 'from(bucket: "{}")\n  |> range(start: {}, stop: {})\n  ' \
                           '|> map(fn: (r) => ({{ actual: r._value, meter_id: r.meter_id, time: r._time }}))' \
        .format(county_code, local_utc(st), local_utc(et))
    print(infludb_device_query)
    device_result = query_api.query(org='nari', query=infludb_device_query)
    for item in device_result:
        for device_info in item.records:
            time_str = datetime.datetime.strftime(device_info['time'].astimezone(pytz.timezone('Asia/Shanghai')),
                                                  '%Y-%m-%d %H:%M')
            if device_info['meter_id'] in device_dict.keys():
                if time_str in device_dict[device_info['meter_id']]['power']:
                    device_dict[device_info['meter_id']]['power'][time_str]['actual'] = device_info['actual']

target_device = []
# 计算准确率
for key in device_dict:
    powerSquare = Decimal(0)
    num = 0
    power = device_dict[key]['power']
    for time_str in power:
        if 'actual' in power[time_str] and 'inst_cap' in device_dict[key]:
            temp_rate = (Decimal(power[time_str]['actual']) * 1000 - Decimal(power[time_str]['forecast'])) / Decimal(
                device_dict[key]['inst_cap'])
            # print(key, time_str, power[time_str]['actual'], Decimal(power[time_str]['actual']) * 1000, Decimal(power[time_str]['forecast']), Decimal(device_dict[key]['inst_cap']))
            # print(temp_rate, temp_rate ** 2)
            powerSquare = temp_rate ** 2 + powerSquare
            num += 1
    if num > 0:
        rate = 1 - math.sqrt(powerSquare / Decimal(num))
        target_device.append([key, device_dict[key]['name'], round(rate * 100, 2)])

target_platform = []

table_header = ('编号', '名称', '准确率（%）')
target = [target_platform, target_device]
output_excel_xlsx("t3.xlsx", table_header, target, ['台区', '电表'])

print(target_device)
