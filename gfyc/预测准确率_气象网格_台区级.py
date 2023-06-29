import argparse
import datetime
import json
import math
import os
import sys
from decimal import *

import influxdb_client
import pymysql
import pytz
import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side

'''
    功能      ->    导出台区级准确率表格（可指定区县，日期）
'''

parser = argparse.ArgumentParser(description='导出台区级准确率表格')
parser.add_argument('--countyCode', type=str, help='区县编号（不填取所有，多个用逗号分隔）')
parser.add_argument('--date', type=str, help='处理日期')

start_time_str = '00:00:00'
end_time_str = '23:59:59'
date_str = datetime.datetime.strftime(datetime.datetime.now() + datetime.timedelta(days=-1), '%Y-%m-%d')

county_code_list = []
# 档案映射
grid_platform_dict = {}
platform_device_dict = {}
county_platform_dict = {}
device_dict = {}
county_dict = {}

# 数据
weather_dict = {}
platform_power_data = {}
device_power_data = {}
total_platform_data = []
accuracy_rate_data = []

conn = pymysql.connect(user='root', password='123456789', host='172.16.130.188', database='gfyc')
cursor = conn.cursor()

client = influxdb_client.InfluxDBClient(url='http://172.16.130.205:8092/',
                                        token='LRdIIW17oir2cDvCsrjZn1qvUOF5fFNwlqeqHGvg5LAv7pw-g_efZNbp7hhvV1aZwBecmMNgeE8dO9yPIPdLqA==',
                                        org='nari',
                                        timeout=50_000)
query_api = client.query_api()


# 本地时间转utc时间
def local_utc(datetime_str):
    local_time = datetime.datetime.strptime(datetime_str, '%Y-%m-%d %H:%M:%S')
    utc_time = local_time.astimezone(pytz.timezone('UTC'))
    utc_time_str = datetime.datetime.strftime(utc_time, '%Y-%m-%dT%H:%M:%SZ')
    return utc_time_str


# 获取区县信息
def init_county_code(county_code):
    global county_code_list
    global county_dict

    if county_code is not None:
        county_code_arr = county_code.split(',')
        county_code_query = 'where county_code in ('
        for index, item in enumerate(county_code_arr):
            if index > 0:
                county_code_query += ', '
            county_code_query += '"' + item + '"'
        county_code_query += ')'
    else:
        county_code_query = ''
    county_code_sql = 'SELECT DISTINCT county_code, county_name FROM dwd_inst_cap {}'.format(county_code_query)
    print('【获取所有区县编码】 \n' + county_code_sql + '\n')
    cursor.execute(county_code_sql)
    county_code_result = cursor.fetchall()
    for item in county_code_result:
        county_code_list.append(item[0])
        if item[0] not in county_dict:
            county_dict[item[0]] = {
                'name': item[1]
            }


# 初始化档案信息
def init_archives_info():
    global county_code_list
    global grid_platform_dict
    global platform_device_dict
    global county_platform_dict
    global device_dict
    # 获取档案信息
    archives_sql = 'SELECT \n' \
                   '    a.county_code, \n' \
                   '    a.county_name, \n' \
                   '    a.line_id, \n' \
                   '    a.line_name, \n' \
                   '    a.resrc_supl_code, \n' \
                   '    a.resrc_supl_name, \n' \
                   '    a.dev_no, \n' \
                   '    a.cust_name, \n' \
                   '    a.weather_grid_id, \n' \
                   '    a.inst_cap, \n' \
                   '    b.per, \n' \
                   '    b.grid_id \n' \
                   'FROM \n' \
                   '    dwd_inst_cap a \n' \
                   '    JOIN pv_device_per_grid_full b ON a.dev_no = b.dev_no \n' \
                   'WHERE \n' \
                   '    a.inst_cap > 0 \n' \
                   '    AND a.weather_grid_id > 0 \n' \
                   '    AND b.grid_id is not null \n'
    if 0 < len(county_code_list) < 30:
        archives_sql += '    AND a.county_code in (' + ','.join(county_code_list) + ')'
    print('【初始化档案信息】 \n' + archives_sql + '\n')
    cursor.execute(archives_sql)
    archives_result = cursor.fetchall()

    for item in archives_result:
        #  台区 - 电表 映射
        if item[4] in platform_device_dict:
            if item[6] not in platform_device_dict[item[4]]['child']:
                platform_device_dict[item[4]]['child'].append(item[6])
            platform_device_dict[item[4]]['inst_cap'] += float(item[9])
        else:
            platform_device_dict[item[4]] = {
                'name': item[5],
                'child': [item[6]],
                'inst_cap': float(item[9])
            }

        # 网格 - 台区 映射
        if item[11] not in grid_platform_dict:
            grid_platform_dict[item[11]] = []
        if item[4] not in grid_platform_dict[item[11]]:
            grid_platform_dict[item[11]].append(item[4])

        # 区县 - 台区 映射
        if item[0] not in county_platform_dict:
            county_platform_dict[item[0]] = []
        if item[4] not in county_platform_dict[item[0]]:
            county_platform_dict[item[0]].append(item[4])

        # 电表信息
        device_dict[item[6]] = {
            'name': item[7],
            'inst_cap': float(item[9]),
            'per': item[10],
        }

    archives_dict = {
        'grid': grid_platform_dict,
        'platform': platform_device_dict,
        'device': device_dict
    }
    with open('档案数据.json', 'w', encoding='utf-8') as target_file:
        json.dump(archives_dict, target_file, indent=4, ensure_ascii=False)
    target_file.close()


# 获取天气
def get_weather():
    weather_county_code = ''
    if len(county_code_list) == 1:
        weather_county_code = '     AND g.county_code = "' + county_code_list[0] + '" \n'
    weather_sql = 'SELECT \n' \
                  '     LEFT(s.yb_time, 10) AS date, \n' \
                  '     g.county_code, \n' \
                  '     CASE \n' \
                  '         s.weather_text \n' \
                  '         WHEN "晴" THEN "晴" \n' \
                  '         WHEN "多云" THEN "多云" \n' \
                  '         ELSE "阴雨" \n' \
                  '     END weather, \n' \
                  '     COUNT( 0 ) num  \n' \
                  'FROM \n' \
                  '     gfyc_relation g, \n' \
                  '     system_forecast_qx_hour_buffer_all s \n' \
                  'WHERE \n' \
                  '     g.station_id = s.id \n' \
                  '{}' \
                  '     AND s.yb_time between "{} {}" and "{} {}" \n' \
                  'GROUP BY date, county_code, weather \n' \
                  'ORDER BY date, county_code, num DESC \n' \
        .format(weather_county_code, date_str, start_time_str, date_str, end_time_str)

    print("【获取天气】 \n" + weather_sql)
    cursor.execute(weather_sql)
    weather_result = cursor.fetchall()
    for item in weather_result:
        if item[0] not in weather_dict:
            weather_dict[item[0]] = {}
        if item[1] not in weather_dict[item[0]]:
            weather_dict[item[0]][item[1]] = item[2]


# 获取预测值
def get_forecast():
    st = datetime.datetime.now()
    global device_power_data
    forecast_sql = 'SELECT \n' \
                   '    grid_id,  \n' \
                   '    time,  \n' \
                   '    FORECAST  \n' \
                   'FROM \n' \
                   '    gfyc_shortforecast_grid \n' \
                   'WHERE \n' \
                   '    time BETWEEN "{} {}" AND "{} {}" ' \
        .format(date_str, start_time_str, date_str, end_time_str)
    print('【获取全量预测值】 开始 \n {}'.format(forecast_sql))
    query_et = datetime.datetime.now()
    cursor.execute(forecast_sql)
    forecast_result = cursor.fetchall()
    for item in forecast_result:
        if item[0] in grid_platform_dict:
            for platform in grid_platform_dict[item[0]]:
                for device in platform_device_dict[platform]['child']:
                    if device not in device_power_data:
                        device_power_data[device] = {}
                    time_str = datetime.datetime.strftime(item[1], '%Y-%m-%d %H:%M')
                    forecast_value = round(item[2] * device_dict[device]['per'], 5)
                    if time_str in device_power_data[device]:
                        device_power_data[device][time_str]['forecast'] = forecast_value
                    else:
                        device_power_data[device][time_str] = {
                            'forecast': forecast_value
                        }
    print('【获取全量预测值】 完成，获取 {} s， 处理 {} s \n'.format((query_et - st).seconds, (datetime.datetime.now() - st).seconds))


# 获取实际值
def get_actual(county_code):
    global device_power_data
    actual_query = 'from(bucket: "{}")\n' \
                   '  |> range(start: {}, stop: {})\n' \
                   '  |> drop(columns: ["_start", "_stop", "_field", "_measurement", "meter_id", "platform_id"])' \
        .format(county_code, local_utc(date_str + ' ' + start_time_str), local_utc(date_str + ' ' + end_time_str))
    print('【获取实际值】 {}  -  {}\n{} \n'.format(date_str, county_dict[county_code]['name'], actual_query))
    actual_result = query_api.query(query=actual_query)
    for item in actual_result:
        for info in item.records:
            if info['device_id'] not in device_power_data:
                device_power_data[info['device_id']] = {}
            time_str = datetime.datetime.strftime(info['_time'].astimezone(pytz.timezone('Asia/Shanghai')),
                                                  '%Y-%m-%d %H:%M')
            actual_value = info['_value'] * 1000
            if time_str in device_power_data[info['device_id']]:
                device_power_data[info['device_id']][time_str]['actual'] = actual_value
            else:
                device_power_data[info['device_id']][time_str] = {
                    'actual': actual_value
                }


# 合并 区县、馈线、网格、台区 数据
def merge_data(county_code):
    global platform_power_data
    platform_power_data = {}

    for platform in county_platform_dict[county_code]:
        if platform in platform_device_dict:
            for child_id in platform_device_dict[platform]['child']:
                if child_id in device_power_data:
                    if platform not in platform_power_data:
                        platform_power_data[platform] = {}
                    for time_str in device_power_data[child_id]:
                        device_power = device_power_data[child_id][time_str]
                        if 'forecast' in device_power and 'actual' in device_power:
                            if time_str not in platform_power_data[platform]:
                                platform_power_data[platform][time_str] = {
                                    'actual': device_power['actual'],
                                    'forecast': device_power['forecast']
                                }
                            else:
                                platform_power_data[platform][time_str]['actual'] += device_power['actual']
                                platform_power_data[platform][time_str]['forecast'] += device_power['forecast']


# 计算准确率
def calculation_percent():
    global accuracy_rate_data
    global total_platform_data
    target_data = []
    sum_rate = Decimal(0)
    for field_id in platform_power_data:
        if field_id in platform_device_dict:
            power_square = Decimal(0)
            num = 0
            for time_str in platform_power_data[field_id]:
                power = platform_power_data[field_id][time_str]
                temp_rate = (Decimal(power['actual']) - Decimal(power['forecast'])) / (
                        Decimal(platform_device_dict[field_id]['inst_cap']) * 1000)
                power_square = temp_rate ** 2 + power_square
                num += 1
            if num > 0:
                cell_data = [field_id, platform_device_dict[field_id]['name'],
                             platform_device_dict[field_id]['inst_cap'], num]
                rate = 1 - math.sqrt(power_square / Decimal(num))
                cell_data.append(round(rate * 100, 2))
                sum_rate = sum_rate + Decimal(rate)
                if cell_data[4] < 0:
                    cell_data.append(1)
                else:
                    cell_data.append(0)
                target_data.append(cell_data)
    # if len(target_data) > 0:
    #     target_data.insert(0, ['平均值', '', '', '',
    #                            round(float(sum_rate) / len(target_data) * 100, 2)])
    total_platform_data.extend(target_data)
    accuracy_rate_data.append(target_data)


# 表头
table_header = ['编号', '名称', '开机容量', '计算点数', '模型准确率（%）', '是否异常(0:正常，1:异常)']
# 表空间
sheet_list = []


# 数据导出到表格
def export_data(target_file, sheet_data, power_data):
    book = openpyxl.Workbook()
    for i in range(0, len(sheet_data)):
        sheet = book.create_sheet(sheet_data[i], i)
        power_data[i].insert(0, table_header)
        # 写数据
        row_num = 1
        for row_data in power_data[i]:
            col_num = 1
            for single_data in row_data:
                sheet.cell(row_num, col_num, single_data)
                col_num += 1
            row_num += 1
        # 修改单元格格式
        fill = PatternFill(patternType='solid', start_color='38B0DE')
        font = Font(bold=True)
        thin = Side(border_style='thin', color='000000')
        border = Border(left=thin, right=thin, bottom=thin, top=thin)
        for i in sheet[1]:
            i.fill = fill
            i.font = font
            i.border = border
            sheet.column_dimensions[i.column_letter].width = 40
    book.save(target_file)


# 获取数据
def get_data():
    st = datetime.datetime.now()
    global device_power_data
    global platform_power_data
    global accuracy_rate_data
    global total_platform_data
    print('【导出开始】 {} \n'.format(date_str))
    # 获取预测值
    get_forecast()
    i = 1
    for county_code in county_code_list:
        print("【区县处理】 {} / {}".format(i, len(county_code_list)))
        i += 1
        # 获取实际值
        get_actual(county_code)
        # 合并数据
        merge_data(county_code)
        # 采集成功率
        calculation_percent()
        weather_str = ''
        if date_str in weather_dict:
            if county_code in weather_dict[date_str]:
                weather_str = weather_dict[date_str][county_code]
        sheet_list.append(county_dict[county_code]['name'] + '（' + weather_str + '）')

    with open('功率电表数据.json', 'w', encoding='utf-8') as target_file:
        json.dump(device_power_data, target_file, indent=4, ensure_ascii=False)
    target_file.close()

    with open('功率台区数据.json', 'w', encoding='utf-8') as target_file:
        json.dump(platform_power_data, target_file, indent=4, ensure_ascii=False)
    target_file.close()

    # 台区分开导出
    export_data('预测准确率-台区拆分({}).xlsx'.format(date_str), sheet_list, accuracy_rate_data)
    # 台区合并导出
    export_data('预测准确率-台区合并({}).xlsx'.format(date_str), ['台区'], [total_platform_data])

    print("【导出完成】 {} ，耗时 {} min\n\n".format(date_str, (datetime.datetime.now() - st).seconds // 60))


if __name__ == '__main__':
    args = parser.parse_args()
    if args.date is not None:
        date_str = args.date
    init_county_code(args.countyCode)
    init_archives_info()
    get_weather()
    # get_data()
