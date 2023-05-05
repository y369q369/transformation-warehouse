import argparse
import datetime
import json
import math
from decimal import *

import influxdb_client
import pymysql
import pytz
import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side

'''
    功能      ->    导出各级别准确率表格（日期）
'''

parser = argparse.ArgumentParser(description='导出台区级准确率表格')
parser.add_argument('--date', type=str, help='导出日期，默认昨日')

start_time_str = '00:00:00'
end_time_str = '23:59:59'

# 导出日期
export_date = ''
# 档案映射
archives_dict = {
    'county_line': {},
    'line_platform': {},
    'grid_platform': {},
    'platform_device': {},
    'device': {},
}

# 数据
weather_dict = {}
archives_power_data = {
    'county': {},
    'line': {},
    'grid': {},
    'platform': {},
    'device': {},
}

conn = pymysql.connect(user='root', password='123456789', host='172.16.130.188', database='gfyc')
cursor = conn.cursor()

client = influxdb_client.InfluxDBClient(url='http://172.16.130.205:8092/',
                                        token='LRdIIW17oir2cDvCsrjZn1qvUOF5fFNwlqeqHGvg5LAv7pw-g_efZNbp7hhvV1aZwBecmMNgeE8dO9yPIPdLqA==',
                                        org='nari',
                                        timeout=50_000)
query_api = client.query_api()


# 本地时间转utc时间
def local_utc(datetime_str):
    local_time = datetime.datetime.strptime(datetime_str, '%Y-%m-%d %H:%M:%S')
    utc_time = local_time.astimezone(pytz.timezone('UTC'))
    utc_time_str = datetime.datetime.strftime(utc_time, '%Y-%m-%dT%H:%M:%SZ')
    return utc_time_str


# 初始化档案信息
def init_archives_info():
    # 获取档案信息
    archives_sql = 'SELECT \n' \
                   '    a.county_code, \n' \
                   '    a.county_name, \n' \
                   '    a.line_id, \n' \
                   '    a.line_name, \n' \
                   '    a.resrc_supl_code, \n' \
                   '    a.resrc_supl_name, \n' \
                   '    a.dev_no, \n' \
                   '    a.cust_name, \n' \
                   '    a.weather_grid_id, \n' \
                   '    a.inst_cap, \n' \
                   '    b.per \n' \
                   'FROM \n' \
                   '    dwd_inst_cap a \n' \
                   '    JOIN pv_device_per_full b ON a.dev_no = b.dev_no \n' \
                   'WHERE \n' \
                   '    a.inst_cap > 0 \n' \
                   '    AND a.weather_grid_id > 0'
    print('【初始化档案信息】 \n' + archives_sql + '\n')
    cursor.execute(archives_sql)
    archives_result = cursor.fetchall()

    for item in archives_result:
        # 区县 - 馈线映射
        if item[0] in archives_dict['county_line']:
            if item[2] not in archives_dict['county_line'][item[0]]['child']:
                archives_dict['county_line'][item[0]]['child'].append(item[2])
            archives_dict['county_line'][item[0]]['inst_cap'] += float(item[9])
        else:
            archives_dict['county_line'][item[0]] = {
                'name': item[1],
                'child': [item[2]],
                'inst_cap': float(item[9])
            }

        # 馈线 - 台区
        if item[2] in archives_dict['line_platform']:
            if item[4] not in archives_dict['line_platform'][item[2]]['child']:
                archives_dict['line_platform'][item[2]]['child'].append(item[4])
            archives_dict['line_platform'][item[2]]['inst_cap'] += float(item[9])
        else:
            archives_dict['line_platform'][item[2]] = {
                'name': item[3],
                'child': [item[4]],
                'inst_cap': float(item[9])
            }

        # 气象网格 - 台区 映射
        if item[8] in archives_dict['grid_platform']:
            if item[4] not in archives_dict['grid_platform'][item[8]]['child']:
                archives_dict['grid_platform'][item[8]]['child'].append(item[4])
            archives_dict['grid_platform'][item[8]]['inst_cap'] += float(item[9])
        else:
            archives_dict['grid_platform'][item[8]] = {
                'name': item[8],
                'child': [item[4]],
                'inst_cap': float(item[9])
            }

        #  台区 - 电表 映射
        if item[4] in archives_dict['platform_device']:
            if item[6] not in archives_dict['platform_device'][item[4]]['child']:
                archives_dict['platform_device'][item[4]]['child'].append(item[6])
            archives_dict['platform_device'][item[4]]['inst_cap'] += float(item[9])
        else:
            archives_dict['platform_device'][item[4]] = {
                'name': item[5],
                'child': [item[6]],
                'inst_cap': float(item[9])
            }

        # 电表信息
        archives_dict['device'][item[6]] = {
            'name': item[7],
            'inst_cap': float(item[9]),
            'per': item[10],
        }


# 获取实际值
def get_actual():
    global archives_power_data
    print('【获取实际值】')
    for county_code in archives_dict['county_line'].keys():
        power = archives_power_data['device']
        actual_query = 'from(bucket: "{}")\n' \
                       '  |> range(start: {}, stop: {})\n' \
                       '  |> drop(columns: ["_start", "_stop", "_field", "_measurement", "meter_id", "platform_id"])' \
            .format(county_code, local_utc(export_date + ' ' + start_time_str),
                    local_utc(export_date + ' ' + end_time_str))
        print('   {} \n {}'.format(archives_dict['county_line'][county_code]['name'], actual_query))
        actual_result = query_api.query(query=actual_query)
        for item in actual_result:
            for info in item.records:
                if info['device_id'] not in power:
                    power[info['device_id']] = {}
                time_str = datetime.datetime.strftime(info['_time'].astimezone(pytz.timezone('Asia/Shanghai')),
                                                      '%Y-%m-%d %H:%M')
                power[info['device_id']][time_str] = {
                    'actual': info['_value'] * 1000
                }


# 获取预测值
def get_forecast():
    global archives_power_data
    print('\n【获取预测值】')
    for county_code in archives_dict['county_line'].keys():
        forecast_sql = 'SELECT \n' \
                       '    a.dev_no,  \n' \
                       '    b.time,  \n' \
                       '    round( a.per * b.FORECAST, 5 )  \n' \
                       'FROM \n' \
                       '    pv_device_per_full a, gfyc_shortforecast b \n' \
                       'WHERE \n' \
                       '    a.county_code = b.id \n' \
                       '    AND b.time BETWEEN "{} {}" AND "{} {}" \n' \
                       '    AND a.county_code = "{}"' \
            .format(export_date, start_time_str, export_date, end_time_str, county_code)
        print('    {} \n {}'.format(archives_dict['county_line'][county_code]['name'], forecast_sql))
        cursor.execute(forecast_sql)
        forecast_result = cursor.fetchall()
        for item in forecast_result:
            if item[0] in archives_power_data['device']:
                time_str = datetime.datetime.strftime(item[1], '%Y-%m-%d %H:%M')
                if time_str in archives_power_data['device'][item[0]]:
                    archives_power_data['device'][item[0]][time_str]['forecast'] = item[2]


# 合并 区县、馈线、网格、台区 数据
def merge_data():
    global archives_power_data
    archives_power_data['county'] = {}
    archives_power_data['line'] = {}
    archives_power_data['grid'] = {}
    archives_power_data['platform'] = {}

    field = ['platform_device', 'grid_platform', 'line_platform', 'county_line']
    power_field = ['platform', 'grid', 'line', 'county']
    base_power_field = ['device', 'platform', 'platform', 'line']

    for i in range(len(field)):
        for self in archives_dict[field[i]]:
            power = archives_power_data[power_field[i]]
            base_power = archives_power_data[base_power_field[i]]

            power[self] = {}
            for child_id in archives_dict[field[i]][self]['child']:
                if child_id in base_power:
                    for time_str in base_power[child_id]:
                        if time_str not in power[self]:
                            power[self][time_str] = {
                                'forecast': base_power[child_id][time_str]['forecast'],
                                'actual': base_power[child_id][time_str]['actual']
                            }
                        else:
                            power[self][time_str]['forecast'] += base_power[child_id][time_str]['forecast']
                            power[self][time_str]['actual'] += base_power[child_id][time_str]['actual']


# 计算准确率
def calculation_percent():
    target_data = []
    power_field = ['county', 'line', 'grid', 'platform', 'device']
    info_field = ['county_line', 'line_platform', 'grid_platform', 'platform_device', 'device']
    for i in range(len(power_field)):
        field_data = []
        for field_id in archives_power_data[power_field[i]]:
            if field_id in archives_dict[info_field[i]]:
                power_square = Decimal(0)
                num = 0
                for time_str in archives_power_data[power_field[i]][field_id]:
                    power = archives_power_data[power_field[i]][field_id][time_str]
                    temp_rate = (Decimal(power['actual']) - Decimal(power['forecast'])) / (
                            Decimal(archives_dict[info_field[i]][field_id]['inst_cap']) * 1000)
                    power_square = temp_rate ** 2 + power_square
                    num += 1
                if num > 0:
                    rate = 1 - math.sqrt(power_square / Decimal(num))
                    field_data.append([field_id, archives_dict[info_field[i]][field_id]['name'],
                                       archives_dict[info_field[i]][field_id]['inst_cap'], num, round(rate * 100, 2)])
        target_data.append(field_data)
    target_file = '预测准确率({}).xlsx'.format(export_date)
    export_data(target_data, target_file)
    print("【导出完成】 " + export_date + "\n\n")


# 表头
table_header = ['编号', '名称', '开机容量', '计算点数', '准确率（%）']
# 表空间
sheet_list = ['区县', '馈线', '网格', '台区', '电表']


# 数据导出到表格
def export_data(target_data, target_file):
    book = openpyxl.Workbook()
    for i in range(0, len(sheet_list)):
        sheet = book.create_sheet(sheet_list[i], i)
        target_data[i].insert(0, table_header)
        # 写数据
        row_num = 1
        for row_data in target_data[i]:
            col_num = 1
            for single_data in row_data:
                sheet.cell(row_num, col_num, single_data)
                col_num += 1
            row_num += 1
        # 修改单元格格式
        fill = PatternFill(patternType='solid', start_color='38B0DE')
        font = Font(bold=True)
        thin = Side(border_style='thin', color='000000')
        border = Border(left=thin, right=thin, bottom=thin, top=thin)
        for i in sheet[1]:
            i.fill = fill
            i.font = font
            i.border = border
            sheet.column_dimensions[i.column_letter].width = 40
    book.save(target_file)


# 获取数据
def get_data():
    global archives_power_data
    # with open('功率电表数据.json', 'r', encoding='utf-8') as load_f:
    #     archives_power_data = json.load(load_f)

    # 获取实际值
    get_actual()
    # # 获取预测数据
    get_forecast()

    # with open('功率电表数据.json', 'w', encoding='utf-8') as target_file:
    #     json.dump(archives_power_data, target_file, indent=4, ensure_ascii=False)
    # target_file.close()

    # 合并数据
    merge_data()

    # with open('功率电表数据2.json', 'w', encoding='utf-8') as target_file:
    #     json.dump(archives_power_data, target_file, indent=4, ensure_ascii=False)
    # target_file.close()

    calculation_percent()


if __name__ == '__main__':
    st = datetime.datetime.now()
    args = parser.parse_args()
    if args.date is None:
        export_date = datetime.datetime.strftime(datetime.datetime.now() + datetime.timedelta(days=-1), '%Y-%m-%d')
    else:
        export_date = args.date
    init_archives_info()
    get_data()
    print(f'耗时 {(datetime.datetime.now() - st).seconds // 60} 分钟')
