import argparse
import datetime
import json
import math
import sys
from decimal import *

import influxdb_client
import pymysql
import pytz
import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side

'''
    功能      ->    多日导出单个表格,统计每日4个模型台区级光伏功率预测准确率
'''

parser = argparse.ArgumentParser(description='多日导出单个表格,统计每日4个模型台区级光伏功率预测准确率')
parser.add_argument('--countyCode', type=str, help='区县编号', default='3241106')
parser.add_argument('--startDate', type=str, help='开始日期')
parser.add_argument('--endDate', type=str, help='结束日期')

start_time_str = '00:00:00'
end_time_str = '23:59:59'

county_code = ''
# 档案映射
platform_device_dict = {}
device_dict = {}

# 数据
date_list = []
weather_dict = {}
platform_power_data = {}
device_power_data = {}

conn = pymysql.connect(user='root', password='123456789', host='172.16.130.188', database='gfyc')
cursor = conn.cursor()

client = influxdb_client.InfluxDBClient(url='http://172.16.130.205:8092/',
                                        token='LRdIIW17oir2cDvCsrjZn1qvUOF5fFNwlqeqHGvg5LAv7pw-g_efZNbp7hhvV1aZwBecmMNgeE8dO9yPIPdLqA==',
                                        org='nari',
                                        timeout=50_000)
query_api = client.query_api()


# 本地时间转utc时间
def local_utc(datetime_str):
    local_time = datetime.datetime.strptime(datetime_str, '%Y-%m-%d %H:%M:%S')
    utc_time = local_time.astimezone(pytz.timezone('UTC'))
    utc_time_str = datetime.datetime.strftime(utc_time, '%Y-%m-%dT%H:%M:%SZ')
    return utc_time_str


# 初始化档案信息
def init_archives_info():
    global platform_device_dict
    global device_dict
    print('【区县编码】 ' + county_code + '\n')
    # 获取档案信息
    archives_sql = 'SELECT \n' \
                   '    a.county_code, \n' \
                   '    a.county_name, \n' \
                   '    a.line_id, \n' \
                   '    a.line_name, \n' \
                   '    a.resrc_supl_code, \n' \
                   '    a.resrc_supl_name, \n' \
                   '    a.meter_id, \n' \
                   '    a.cust_name, \n' \
                   '    a.weather_grid_id, \n' \
                   '    a.inst_cap, \n' \
                   '    b.per \n' \
                   'FROM \n' \
                   '    dwd_inst_cap a \n' \
                   '    JOIN pv_device_per_full b ON a.dev_no = b.dev_no \n' \
                   'WHERE \n' \
                   '    a.county_code = "{}" \n' \
                   '    AND a.inst_cap > 0 \n' \
                   '    AND a.weather_grid_id >0'.format(county_code)
    print('【初始化档案信息】 \n' + archives_sql + '\n')
    cursor.execute(archives_sql)
    archives_result = cursor.fetchall()

    for item in archives_result:
        #  台区 - 电表 映射
        if item[4] in platform_device_dict:
            if item[6] not in platform_device_dict[item[4]]['child']:
                platform_device_dict[item[4]]['child'].append(item[6])
            platform_device_dict[item[4]]['inst_cap'] += float(item[9])
        else:
            platform_device_dict[item[4]] = {
                'name': item[5],
                'child': [item[6]],
                'inst_cap': float(item[9])
            }

        # 电表信息
        device_dict[item[6]] = {
            'name': item[7],
            'inst_cap': float(item[9]),
            'per': item[10],
        }


# 获取天气
def get_weather():
    print("【获取天气】")
    for date_str in date_list:
        weather_sql = 'SELECT \n' \
                      '     CASE \n' \
                      '         s.weather_text \n' \
                      '         WHEN "晴" THEN "晴" \n' \
                      '         WHEN "多云" THEN "多云" \n' \
                      '         ELSE "阴雨" \n' \
                      '     END weather, \n' \
                      '     COUNT( 0 ) num  \n' \
                      'FROM \n' \
                      '     gfyc_relation g, \n' \
                      '     system_forecast_qx_hour_buffer_all s \n' \
                      'WHERE \n' \
                      '     g.station_id = s.id \n' \
                      '     AND g.county_code = "{}" \n' \
                      '     AND s.yb_time between "{} {}" and "{} {}" \n' \
                      'GROUP BY weather \n' \
                      'ORDER BY num DESC \n' \
                      'LIMIT 1 \n'.format(county_code, date_str, start_time_str, date_str, end_time_str)
        print(weather_sql)
        cursor.execute(weather_sql)
        inst_cap_result = cursor.fetchone()
        weather_dict[date_str] = inst_cap_result[0]


# 获取实际值
def get_actual(date_str):
    print("【获取实际值】")
    global device_power_data
    device_power_data = {}
    actual_query = 'import "date"\n' \
                   'from(bucket: "{}")\n' \
                   '  |> range(start: {}, stop: {})\n' \
                   '  |> filter(fn:(r) => date.minute(t: r._time) == 0) \n' \
                   '  |> drop(columns: ["_start", "_stop", "_field", "_measurement", "device_id", "platform_id"])' \
        .format(county_code, local_utc(date_str + ' ' + start_time_str), local_utc(date_str + ' ' + end_time_str))
    print(actual_query + '\n')
    actual_result = query_api.query(query=actual_query)
    for item in actual_result:
        for info in item.records:
            if info['meter_id'] not in device_power_data:
                device_power_data[info['meter_id']] = {}
            time_str = datetime.datetime.strftime(info['_time'].astimezone(pytz.timezone('Asia/Shanghai')),
                                                  '%Y-%m-%d %H:%M')
            device_power_data[info['meter_id']][time_str] = {
                'actual': info['_value'] * 1000
            }


# 获取预测值
def get_forecast(date_str):
    for i in range(1, 5):
        print("【获取预测值】 模型" + str(i))
        forecast_sql = 'SELECT \n' \
                       '    a.meter_id,  \n' \
                       '    b.time,  \n' \
                       '    round( a.per * b.FORECAST, 5 )  \n' \
                       'FROM \n' \
                       '    pv_device_per_full a, gfyc_shortforecast_{} b \n' \
                       'WHERE \n' \
                       '    a.county_code = b.id \n' \
                       '    AND b.time BETWEEN "{} {}" AND "{} {}" \n' \
                       '    AND a.county_code = "{}" ' \
            .format(i, date_str, start_time_str, date_str, end_time_str, county_code)
        print(forecast_sql + '\n')
        cursor.execute(forecast_sql)
        forecast_result = cursor.fetchall()
        for item in forecast_result:
            if item[0] in device_power_data:
                time_str = datetime.datetime.strftime(item[1], '%Y-%m-%d %H:%M')
                if time_str in device_power_data[item[0]]:
                    device_power_data[item[0]][time_str]['forecast' + str(i)] = item[2]


# 合并 区县、馈线、网格、台区 数据
def merge_data():
    global platform_power_data
    platform_power_data = {}

    for self in platform_device_dict:
        for child_id in platform_device_dict[self]['child']:
            if child_id in device_power_data:
                if self not in platform_power_data:
                    platform_power_data[self] = {}
                for time_str in device_power_data[child_id]:
                    if time_str not in platform_power_data[self]:
                        platform_power_data[self][time_str] = {
                            'actual': device_power_data[child_id][time_str]['actual']
                        }
                        for i in range(1, 5):
                            platform_power_data[self][time_str]['forecast' + str(i)] = \
                            device_power_data[child_id][time_str]['forecast' + str(i)]
                    else:
                        for i in range(1, 5):
                            platform_power_data[self][time_str]['forecast' + str(i)] += \
                            device_power_data[child_id][time_str]['forecast' + str(i)]
                        platform_power_data[self][time_str]['actual'] += device_power_data[child_id][time_str]['actual']
    with open('功率数据.json', 'w', encoding='utf-8') as target_file:
        json.dump(platform_power_data, target_file, indent=4, ensure_ascii=False)
        target_file.close()
    with open('档案数据.json', 'w', encoding='utf-8') as target_file:
        json.dump(platform_device_dict, target_file, indent=4, ensure_ascii=False)
        target_file.close()


# 计算准确率
def calculation_percent(date_str):
    target_data = []
    sum_rate = {
        '1': Decimal(0),
        '2': Decimal(0),
        '3': Decimal(0),
        '4': Decimal(0)
    }
    for field_id in platform_power_data:
        if field_id in platform_device_dict:
            power_square = {
                '1': Decimal(0),
                '2': Decimal(0),
                '3': Decimal(0),
                '4': Decimal(0),
            }
            num = 0
            for time_str in platform_power_data[field_id]:
                power = platform_power_data[field_id][time_str]
                for i in range(1, 5):
                    temp_rate = (Decimal(power['actual']) - Decimal(power['forecast' + str(i)])) / (
                            Decimal(platform_device_dict[field_id]['inst_cap']) * 1000)
                    power_square[str(i)] = temp_rate ** 2 + power_square[str(i)]
                num += 1
            if num > 0:
                cell_data = [field_id, platform_device_dict[field_id]['name'],
                             platform_device_dict[field_id]['inst_cap'], num]
                for i in range(1, 5):
                    rate = 1 - math.sqrt(power_square[str(i)] / Decimal(num))
                    cell_data.append(round(rate * 100, 2))
                    sum_rate[str(i)] = sum_rate[str(i)] + Decimal(rate)
                if cell_data[4] < 0:
                    cell_data.append(1)
                else:
                    cell_data.append(0)
                target_data.append(cell_data)
    if len(target_data) > 0:
        target_data.insert(0, ['平均值', weather_dict[date_str], '', '',
                               round(float(sum_rate['1']) / len(target_data) * 100, 2),
                               round(float(sum_rate['2']) / len(target_data) * 100, 2),
                               round(float(sum_rate['3']) / len(target_data) * 100, 2),
                               round(float(sum_rate['4']) / len(target_data) * 100, 2)])

    target_file = '预测准确率({}).xlsx'.format(date_str)
    export_data(target_data, target_file)
    print("【导出完成】 " + date_str + "\n\n")


# 表头
table_header = ['编号', '名称', '开机容量', '计算点数', '模型1准确率（%）', '模型2准确率（%）', '模型3准确率（%）', '模型4准确率（%）', '是否异常(0:正常，1:异常)']
# 表空间
sheet_list = ['台区']


# 数据导出到表格
def export_data(target_data, target_file):
    book = openpyxl.Workbook()
    for i in range(0, len(sheet_list)):
        sheet = book.create_sheet(sheet_list[i], i)
        target_data.insert(0, table_header)
        # 写数据
        row_num = 1
        for row_data in target_data:
            col_num = 1
            for single_data in row_data:
                sheet.cell(row_num, col_num, single_data)
                col_num += 1
            row_num += 1
        # 修改单元格格式
        fill = PatternFill(patternType='solid', start_color='38B0DE')
        font = Font(bold=True)
        thin = Side(border_style='thin', color='000000')
        border = Border(left=thin, right=thin, bottom=thin, top=thin)
        for i in sheet[1]:
            i.fill = fill
            i.font = font
            i.border = border
            sheet.column_dimensions[i.column_letter].width = 40
    book.save(target_file)


# 获取数据
def get_data():
    get_weather()
    for date_str in date_list:
        # 获取实际值
        get_actual(date_str)
        # 获取预测数据
        get_forecast(date_str)
        # 合并数据
        merge_data()
        calculation_percent(date_str)


if __name__ == '__main__':
    args = parser.parse_args()
    print(args)
    if args.startDate is None or args.endDate is None:
        date_list.append(datetime.datetime.strftime(datetime.datetime.now() + datetime.timedelta(days=-1), '%Y-%m-%d'))
    else:
        startDate = datetime.datetime.strptime(args.startDate, '%Y-%m-%d')
        endDate = datetime.datetime.strptime(args.endDate, '%Y-%m-%d')
        while endDate >= startDate:
            date_list.append(datetime.datetime.strftime(startDate, '%Y-%m-%d'))
            startDate = startDate + datetime.timedelta(days=1)
    county_code = args.countyCode
    init_archives_info()
    get_data()
