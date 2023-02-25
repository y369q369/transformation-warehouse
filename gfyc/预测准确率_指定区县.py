import argparse
import datetime
import json
import math
import sys
from decimal import *

import influxdb_client
import pymysql
import pytz
import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side

'''
    功能      ->    统计每日光伏功率预测准确率
'''

parser = argparse.ArgumentParser(description='统计每日光伏功率预测准确率，默认查询昨天')
parser.add_argument('--countyCode', type=str, help='区县编号', default='3240107')
parser.add_argument('--startDate', type=str, help='开始日期')
parser.add_argument('--endDate', type=str, help='结束日期')

start_time_str = '08:00:00'
end_time_str = '17:00:00'

county_code = ''
# 档案映射
archives_dict = {
    'county_line': {},
    'line_platform': {},
    'grid_platform': {},
    'platform_device': {},
    'device': {},
}
# 数据
date_list = []
weather_dict = {}
archives_power_data = {
    'county': {},
    'line': {},
    'grid': {},
    'platform': {},
}

client = influxdb_client.InfluxDBClient(url='http://192.168.110.130:8086/',
                                        token='VsE9zz62qm_DZKB6eib-hTSk1Ml-e8uF82Sqdbe5xnUJwOKshHFZCdMaiMa7Fqx_KD2iKmWCjg2u6XHTABcaSA==',
                                        org='gs',
                                        timeout=50_000)
query_api = client.query_api()

conn = pymysql.connect(user='root', password='1qaz@WSX', host='127.0.0.1', database='gfyc')
cursor = conn.cursor()


# 本地时间转utc时间
def local_utc(datetime_str):
    local_time = datetime.datetime.strptime(datetime_str, '%Y-%m-%d %H:%M:%S')
    utc_time = local_time.astimezone(pytz.timezone('UTC'))
    utc_time_str = datetime.datetime.strftime(utc_time, '%Y-%m-%dT%H:%M:%SZ')
    return utc_time_str


# 初始化档案信息
def init_archives_info():
    global archives_dict
    print('【区县编码】 ' + county_code + '\n')
    # 获取档案信息
    archives_sql = 'SELECT \n' \
                   '    a.county_code, \n' \
                   '    a.county_name, \n' \
                   '    a.line_id, \n' \
                   '    a.line_name, \n' \
                   '    a.resrc_supl_code, \n' \
                   '    a.resrc_supl_name, \n' \
                   '    a.meter_id, \n' \
                   '    a.cust_name, \n' \
                   '    a.weather_grid_id, \n' \
                   '    a.inst_cap, \n' \
                   '    b.per \n' \
                   'FROM \n' \
                   '    dwd_inst_cap a \n' \
                   '    JOIN pv_device_per_full b ON a.dev_no = b.dev_no \n' \
                   'WHERE \n' \
                   '    a.county_code = "{}" \n' \
                   '    AND a.inst_cap > 0 \n' \
                   '    AND a.weather_grid_id >0'.format(county_code)
    print('【初始化档案信息】 \n' + archives_sql + '\n')
    cursor.execute(archives_sql)
    archives_result = cursor.fetchall()

    for item in archives_result:
        # 区县 - 馈线映射
        if item[0] in archives_dict['county_line']:
            if item[2] not in archives_dict['county_line'][item[0]]['child']:
                archives_dict['county_line'][item[0]]['child'].append(item[2])
            archives_dict['county_line'][item[0]]['inst_cap'] += float(item[9])
        else:
            archives_dict['county_line'][item[0]] = {
                'name': item[1],
                'child': [item[2]],
                'inst_cap': float(item[9])
            }

        # 馈线 - 台区
        if item[2] in archives_dict['line_platform']:
            if item[4] not in archives_dict['line_platform'][item[2]]['child']:
                archives_dict['line_platform'][item[2]]['child'].append(item[4])
            archives_dict['line_platform'][item[2]]['inst_cap'] += float(item[9])
        else:
            archives_dict['line_platform'][item[2]] = {
                'name': item[3],
                'child': [item[4]],
                'inst_cap': float(item[9])
            }

        # 气象网格 - 台区 映射
        if item[8] in archives_dict['grid_platform']:
            if item[4] not in archives_dict['grid_platform'][item[8]]['child']:
                archives_dict['grid_platform'][item[8]]['child'].append(item[4])
            archives_dict['grid_platform'][item[8]]['inst_cap'] += float(item[9])
        else:
            archives_dict['grid_platform'][item[8]] = {
                'name': item[8],
                'child': [item[4]],
                'inst_cap': float(item[9])
            }

        #  台区 - 电表 映射
        if item[4] in archives_dict['platform_device']:
            if item[6] not in archives_dict['platform_device'][item[4]]['child']:
                archives_dict['platform_device'][item[4]]['child'].append(item[6])
            archives_dict['platform_device'][item[4]]['inst_cap'] += float(item[9])
        else:
            archives_dict['platform_device'][item[4]] = {
                'name': item[5],
                'child': [item[6]],
                'inst_cap': float(item[9])
            }

        # 电表信息
        archives_dict['device'][item[6]] = {
            'name': item[7],
            'inst_cap': float(item[9]),
            'per': item[10],
        }


# 获取天气
def get_weather():
    print("【获取天气】")
    for date_str in date_list:
        weather_sql = 'SELECT \n' \
                      '     CASE \n' \
                      '         s.weather_text \n' \
                      '         WHEN "晴" THEN "晴" \n' \
                      '         WHEN "多云" THEN "多云" \n' \
                      '         ELSE "阴雨" \n' \
                      '     END weather, \n' \
                      '     COUNT( 0 ) num  \n' \
                      'FROM \n' \
                      '     gfyc_relation g, \n' \
                      '     system_forecast_qx_hour_buffer_all s \n' \
                      'WHERE \n' \
                      '     g.station_id = s.id \n' \
                      '     AND g.county_code = "{}" \n' \
                      '     AND s.yb_time between "{} {}" and "{} {}" \n' \
                      'GROUP BY weather \n' \
                      'ORDER BY num DESC \n' \
                      'LIMIT 1 \n'.format(county_code, date_str, start_time_str, date_str, end_time_str)
        print(weather_sql)
        cursor.execute(weather_sql)
        inst_cap_result = cursor.fetchone()
        weather_dict[date_str] = inst_cap_result[0]


# 获取实际值
def get_actual(date_str):
    print("【获取实际值】")
    power = {}
    archives_power_data['device'] = power
    actual_query = 'from(bucket: "{}")\n' \
                   '  |> range(start: {}, stop: {})\n' \
                   '  |> drop(columns: ["_start", "_stop", "_field", "_measurement", "device_id", "platform_id"])' \
        .format(county_code, local_utc(date_str + ' ' + start_time_str), local_utc(date_str + ' ' + end_time_str))
    print(actual_query + '\n')
    actual_result = query_api.query(query=actual_query)
    for item in actual_result:
        for info in item.records:
            if info['meter_id'] not in power:
                power[info['meter_id']] = {}
            time_str = datetime.datetime.strftime(info['_time'].astimezone(pytz.timezone('Asia/Shanghai')),
                                                  '%Y-%m-%d %H:%M')
            power[info['meter_id']][time_str] = {
                'actual': info['_value'] * 1000
            }


# 获取预测值
def get_forecast(date_str):
    print("【获取预测值】 ")
    forecast_sql = 'SELECT \n' \
                   '    a.meter_id,  \n' \
                   '    b.time,  \n' \
                   '    round( a.per * b.FORECAST, 5 )  \n' \
                   'FROM \n' \
                   '    pv_device_per_full a, gfyc_shortforecast b \n' \
                   'WHERE \n' \
                   '    a.county_code = b.id \n' \
                   '    AND b.time BETWEEN "{} {}" AND "{} {}" \n' \
                   '    AND a.county_code = "{}" ' \
        .format(date_str, start_time_str, date_str, end_time_str, county_code)
    print(forecast_sql + '\n')
    cursor.execute(forecast_sql)
    forecast_result = cursor.fetchall()
    for item in forecast_result:
        if item[0] in archives_power_data['device']:
            time_str = datetime.datetime.strftime(item[1], '%Y-%m-%d %H:%M')
            if time_str in archives_power_data['device'][item[0]]:
                archives_power_data['device'][item[0]][time_str]['forecast'] = item[2]


# 合并 区县、馈线、网格、台区 数据
def merge_data():
    global archives_power_data
    archives_power_data['county'] = {}
    archives_power_data['line'] = {}
    archives_power_data['grid'] = {}
    archives_power_data['platform'] = {}

    field = ['platform_device', 'grid_platform', 'line_platform', 'county_line']
    power_field = ['platform', 'grid', 'line', 'county']
    base_power_field = ['device', 'platform', 'platform', 'line']

    for i in range(len(field)):
        for self in archives_dict[field[i]]:
            power = archives_power_data[power_field[i]]
            base_power = archives_power_data[base_power_field[i]]

            power[self] = {}
            for child_id in archives_dict[field[i]][self]['child']:
                if child_id in base_power:
                    for time_str in base_power[child_id]:
                        if time_str not in power[self]:
                            power[self][time_str] = {
                                'forecast': base_power[child_id][time_str]['forecast'],
                                'actual': base_power[child_id][time_str]['actual']
                            }
                        else:
                            power[self][time_str]['forecast'] += base_power[child_id][time_str]['forecast']
                            power[self][time_str]['actual'] += base_power[child_id][time_str]['actual']
    # with open('功率数据.json', 'w', encoding='utf-8') as target_file:
    #     json.dump(archives_power_data, target_file, indent=4, ensure_ascii=False)
    #     target_file.close()
    # with open('档案数据.json', 'w', encoding='utf-8') as target_file:
    #     json.dump(archives_dict, target_file, indent=4, ensure_ascii=False)
    #     target_file.close()


# 计算准确率
def calculation_percent(date_str):
    target_data = []
    power_field = ['county', 'line', 'grid', 'platform', 'device']
    info_field = ['county_line', 'line_platform', 'grid_platform', 'platform_device', 'device']
    for i in range(len(power_field)):
        field_data = []
        for field_id in archives_power_data[power_field[i]]:
            if field_id in archives_dict[info_field[i]]:
                power_square = Decimal(0)
                num = 0
                for time_str in archives_power_data[power_field[i]][field_id]:
                    power = archives_power_data[power_field[i]][field_id][time_str]
                    temp_rate = (Decimal(power['actual']) - Decimal(power['forecast'])) / (
                                Decimal(archives_dict[info_field[i]][field_id]['inst_cap']) * 1000)
                    power_square = temp_rate ** 2 + power_square
                    num += 1
                if num > 0:
                    rate = 1 - math.sqrt(power_square / Decimal(num))
                    field_data.append([field_id, archives_dict[info_field[i]][field_id]['name'],
                                       archives_dict[info_field[i]][field_id]['inst_cap'], num, round(rate * 100, 2)])
        target_data.append(field_data)

    target_file = '预测准确率({})({}).xlsx'.format(date_str, weather_dict[date_str])
    export_data(target_data, target_file)
    print("【导出完成】 " + date_str + "\n\n")


# 表头
table_header = ['编号', '名称', '开机容量', '计算点数', '准确率（%）']
# 表空间
sheet_list = ['区县', '馈线', '网格', '台区', '电表']


# 数据导出到表格
def export_data(target_data, target_file):
    book = openpyxl.Workbook()
    for i in range(0, len(sheet_list)):
        sheet = book.create_sheet(sheet_list[i], i)
        target_data[i].insert(0, table_header)
        # 写数据
        row_num = 1
        for row_data in target_data[i]:
            col_num = 1
            for single_data in row_data:
                sheet.cell(row_num, col_num, single_data)
                col_num += 1
            row_num += 1
        # 修改单元格格式
        fill = PatternFill(patternType='solid', start_color='38B0DE')
        font = Font(bold=True)
        thin = Side(border_style='thin', color='000000')
        border = Border(left=thin, right=thin, bottom=thin, top=thin)
        for i in sheet[1]:
            i.fill = fill
            i.font = font
            i.border = border
            sheet.column_dimensions[i.column_letter].width = 40
    book.save(target_file)


# 获取数据
def get_data():
    get_weather()
    for date_str in date_list:
        # 获取实际值
        get_actual(date_str)
        # 获取预测数据
        get_forecast(date_str)
        # 合并数据
        merge_data()
        calculation_percent(date_str)


if __name__ == '__main__':
    args = parser.parse_args()
    print(args)
    if args.startDate is None or args.endDate is None:
        date_list.append(datetime.datetime.strftime(datetime.datetime.now() + datetime.timedelta(days=-1), '%Y-%m-%d'))
    else:
        startDate = datetime.datetime.strptime(args.startDate, '%Y-%m-%d')
        endDate = datetime.datetime.strptime(args.endDate, '%Y-%m-%d')
        while endDate >= startDate:
            date_list.append(datetime.datetime.strftime(startDate, '%Y-%m-%d'))
            startDate = startDate + datetime.timedelta(days=1)
    county_code = args.countyCode
    init_archives_info()
    get_data()
