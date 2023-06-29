import xlwt
import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side


def sql_output_excel(filename, sheet_name, excel_format="xlsx", sum_title=None):
    # cursor2.execute(sql)

    test = {
        '8200000022600603': {
            'name': '查从兴2',
            'accs_cap': 12000,
            'rate': 0.243
        },
        '8200000022600604': {
            'name': '查从兴',
            'accs_cap': 12040,
            'rate': 0.945
        },
        '8200000022600605': {
            'name': '查从兴3',
            'accs_cap': 12060,
            'rate': 0.9956
        }

    }

    title = ('编号', '用户号', '准确率')
    data = []
    index = 0
    for key in test:
        t = (key, test[key]['name'], test[key]['rate'])
        data.append(t)
        index += 1

    if excel_format == 'xls':
        output_excel_xls(title, data, filename, sheet_name, sum_title=sum_title)
    elif excel_format == 'xlsx':
        output_excel_xlsx(title, data, filename, sheet_name, sum_title=sum_title)


def output_excel_xls(title, data, filename, sheet_name, sum_title=None):
    book = xlwt.Workbook(encoding='utf-8')

    # 设置颜色
    pattern = xlwt.Pattern()
    pattern.pattern = xlwt.Pattern.SOLID_PATTERN
    pattern.pattern_fore_colour = 22

    # 设置边框
    borders = xlwt.Borders()
    borders.right = xlwt.Borders.THIN
    borders.top = xlwt.Borders.THIN
    borders.bottom = xlwt.Borders.THIN
    borders.left = xlwt.Borders.THIN

    style = xlwt.XFStyle()
    style.pattern = pattern
    style.borders = borders

    sheet1 = book.add_sheet(sheet_name)

    # 写入头部
    col_num = 0
    for single_data in title:
        sheet1.write(0, col_num, single_data, style)
        col_num += 1

    # 写数据
    row_num = 1
    for row_data in data:
        col_num = 0
        for single_data in row_data:
            if single_data is None:
                single_data = ""
            sheet1.write(row_num, col_num, str(single_data))
            col_num += 1
        row_num += 1

    book.save(filename)


def output_excel_xlsx(title, data, filename, sheet_name, sum_title=None):
    book = openpyxl.Workbook()
    sheet1 = book.active
    sheet1.title = sheet_name

    data.insert(0, title)

    # 写数据
    row_num = 1
    for row_data in data:
        col_num = 1
        print('-------')
        print(row_data)
        for single_data in row_data:
            sheet1.cell(row_num, col_num, single_data)
            col_num += 1
        row_num += 1

    max_row = sheet1.max_row + 1

    # 修改单元格格式
    fill = PatternFill(patternType='solid', start_color='38B0DE')
    font = Font(bold=True)
    thin = Side(border_style='thin', color='000000')
    border = Border(left=thin, right=thin, bottom=thin, top=thin)
    for i in sheet1[1]:
        i.fill = fill
        i.font = font
        i.border = border

    # 修改时间字段宽度
    for i in sheet1[1]:
        if '时间' in i.value or 'time' in i.value:
            sheet1.column_dimensions[i.column_letter].width = 20

    # 求和
    if sum_title and isinstance(sum_title, list):
        for i in sheet1[1]:
            if i.value in sum_title:
                sum_column = list(sheet1[i.column_letter])
                sum_column.pop(0)

                sum_num = 0
                for single_data in sum_column:
                    if single_data.value:
                        try:
                            sum_num += int(single_data.value)
                        except Exception as e:
                            print(e)
                            break
                else:
                    sheet1.cell(max_row, i.column, sum_num)
    # 添加第二个表空间
    book.create_sheet('台区')
    book.save(filename)


if __name__ == '__main__':
    sql_output_excel("t2.xlsx", "电表", excel_format="xlsx")
